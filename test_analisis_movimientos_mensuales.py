import tempfile
import unittest
from pathlib import Path

import openpyxl
import pandas as pd

import analisis_movimientos_mensuales as analisis


COLUMNAS_SALIDA = [
    "Mes",
    "Ventas",
    "F_Ventas",
    "Nota_c_v",
    "Compras",
    "F_Compras",
    "Nota_c_c",
    "Resultado",
    "% C/V",
    "IIBB",
]


def datos_emitidos():
    return pd.DataFrame(
        {
            "Fecha": ["15/01/2025", "10/02/2025", "05/03/2025"],
            "Tipo de Comprobante": ["1 - Factura A", "3 - Nota de Crédito A", "1 - Factura A"],
            "Número Desde": [1, 2, 3],
            "Denominación Receptor": ["Cliente Grande", "Cliente NC", None],
            "Nro. Doc. Receptor": [30_123_456_789, 30_987_654_321, 99],
            "Imp. Total": ["600.000,00", "100.000,00", "10,00"],
            "Moneda": ["ARS", "ARS", "USD"],
        }
    )


def datos_recibidos():
    return pd.DataFrame(
        {
            "Fecha": ["20/01/2025", "12/02/2025", "08/04/2025"],
            "Tipo": ["1 - Factura A", "8 - Nota de Crédito B", "1 - Factura A"],
            "Número Desde": [10, 11, 12],
            "Denominación Emisor": ["Proveedor Uno", "Proveedor NC", "Proveedor Grande"],
            "Nro. Doc. Emisor": [30_111_111_111, 30_222_222_222, 30_333_333_333],
            "Imp. Total": ["200000", "50.000,00", "700000"],
            "Moneda": ["ARS", "ARS", "ARS"],
        }
    )


class AnalisisAFIPTests(unittest.TestCase):
    def test_deteccion_por_prefijo_y_error_si_hay_duplicados(self):
        with tempfile.TemporaryDirectory() as temporal:
            temporal = Path(temporal)
            nombres = [
                "comprobantes_consulta_csv_emitidos_123_456_20260624.csv",
                "Mis Comprobantes Emitidos (1).xlsx",
                "Mis Comprobantes Recibidos - CUIT 20300412298 (1).xlsx",
            ]
            for nombre in nombres:
                (temporal / nombre).touch()

            encontrados = analisis.listar_archivos_por_tipo(temporal)
            self.assertEqual(len(encontrados["emitidos"]), 2)
            self.assertEqual(len(encontrados["recibidos"]), 1)
            self.assertEqual(
                analisis.clasificar_archivo(temporal / nombres[0]), "emitidos"
            )
            self.assertEqual(
                analisis.clasificar_archivo(temporal / nombres[2]), "recibidos"
            )
            with self.assertRaisesRegex(
                analisis.ErrorDatosAFIP, "más de un archivo del tipo emitidos"
            ):
                analisis.buscar_archivo("emitidos", temporal)

    def test_lectura_csv_y_calculos(self):
        with tempfile.TemporaryDirectory() as temporal:
            ruta = (
                Path(temporal)
                / "comprobantes_consulta_csv_emitidos_123_456_20260624.csv"
            )
            datos = datos_emitidos().rename(columns={"Fecha": "Fecha de Emisión"})
            datos["Fecha de Emisión"] = ["2025-01-15", "2025-02-10", "2025-03-05"]
            ruta.write_text(
                "Reporte exportado desde AFIP\n"
                + datos.to_csv(index=False, sep=";"),
                encoding="utf-8-sig",
            )

            emitidos, columna_tipo = analisis.preparar_comprobantes(
                analisis.leer_archivo_afip(ruta), ruta.name
            )
            resumen = analisis.resumir_por_mes(emitidos, "emitidos")
            grandes = analisis.filtrar_facturas_grandes(emitidos, columna_tipo, "emitidos")

            self.assertEqual(columna_tipo, "Tipo de Comprobante")
            self.assertEqual(emitidos.loc[1, "Imp. Total"], -100_000)
            self.assertEqual(resumen.loc[resumen["Mes"].astype(str).eq("2025-02"), "Ventas"].iat[0], -100_000)
            self.assertEqual(grandes["Operacion"].tolist(), ["Venta", "Venta"])
            self.assertEqual(grandes.loc[1, "CUIT"], "0")
            self.assertEqual(grandes.loc[1, "Denominacion"], "(Consumidor Final)")

    def test_un_solo_tipo_genera_su_resumen_sin_exigir_el_otro(self):
        with tempfile.TemporaryDirectory() as temporal:
            temporal = Path(temporal)
            ruta = temporal / "Mis Comprobantes Recibidos - CUIT 20300412298.xlsx"
            with pd.ExcelWriter(ruta, engine="xlsxwriter") as writer:
                datos_recibidos().to_excel(writer, index=False, startrow=1)
                writer.sheets["Sheet1"].write(0, 0, "Mis Comprobantes Recibidos")

            salida = temporal / "resumen_recibidos.xlsx"
            resultado = analisis.ejecutar_resumen_individual(
                "recibidos", temporal, salida
            )
            libro = openpyxl.load_workbook(resultado, data_only=False)
            self.assertEqual(
                libro.sheetnames, ["Resumen Recibidos", "Facturas Grandes"]
            )
            self.assertEqual(len(libro["Resumen Recibidos"]._charts), 1)
            with self.assertRaisesRegex(
                analisis.ErrorDatosAFIP, "No se encontró el archivo.*emitidos"
            ):
                analisis.ejecutar_analisis(temporal, temporal / "total.xlsx")

    def test_lectura_xlsx_resumen_y_excel_final(self):
        with tempfile.TemporaryDirectory() as temporal:
            temporal = Path(temporal)
            ruta_emitidos = temporal / "Mis Comprobantes Emitidos (1).xlsx"
            ruta_recibidos = temporal / "Mis Comprobantes Recibidos.xlsx"

            for ruta, datos in (
                (ruta_emitidos, datos_emitidos()),
                (ruta_recibidos, datos_recibidos()),
            ):
                with pd.ExcelWriter(ruta, engine="xlsxwriter") as writer:
                    datos.to_excel(writer, index=False, startrow=1)
                    writer.sheets["Sheet1"].write(0, 0, "Reporte exportado desde AFIP")

            emitidos, tipo_emitidos = analisis.preparar_comprobantes(
                analisis.leer_archivo_afip(ruta_emitidos), ruta_emitidos.name
            )
            recibidos, tipo_recibidos = analisis.preparar_comprobantes(
                analisis.leer_archivo_afip(ruta_recibidos), ruta_recibidos.name
            )
            resumen_emitidos = analisis.resumir_por_mes(emitidos, "emitidos")
            resumen_recibidos = analisis.resumir_por_mes(recibidos, "recibidos")
            movimientos = analisis.combinar_resumenes(resumen_emitidos, resumen_recibidos)
            facturas = pd.concat(
                [
                    analisis.filtrar_facturas_grandes(emitidos, tipo_emitidos, "emitidos"),
                    analisis.filtrar_facturas_grandes(recibidos, tipo_recibidos, "recibidos"),
                ],
                ignore_index=True,
            )

            self.assertEqual(movimientos.columns.tolist(), COLUMNAS_SALIDA)
            self.assertEqual(movimientos["Mes"].astype(str).tolist(), ["2025-01", "2025-02", "2025-03", "2025-04"])
            self.assertEqual(movimientos.loc[1, "Resultado"], -50_000)
            self.assertEqual(movimientos.loc[2, "F_Compras"], 0)
            self.assertEqual(facturas["Operacion"].tolist(), ["Venta", "Venta", "Compra"])

            salida = temporal / "resultado.xlsx"
            analisis.generar_excel(movimientos, facturas, salida)
            libro = openpyxl.load_workbook(salida, data_only=False)
            self.assertEqual(libro.sheetnames, ["Resumen Mensual", "Facturas Grandes"])
            self.assertEqual(len(libro["Resumen Mensual"]._charts), 2)
            self.assertEqual(
                [celda.value for celda in libro["Resumen Mensual"][1]],
                COLUMNAS_SALIDA,
            )


if __name__ == "__main__":
    unittest.main()
