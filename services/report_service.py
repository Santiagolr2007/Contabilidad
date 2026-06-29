from __future__ import annotations

import re
from datetime import date, datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from utils.formatters import display_date, display_period, number_ar

from .voucher_service import VoucherService
from .platform_service import PlatformService


class ReportService:
    """Consultas y exportaciones Excel del sistema contable."""

    REPORTS = {
        "ultimos_12_meses": "Últimos 12 meses - Ventas, compras e IIBB",
        "ventas_mensuales": "Ventas mensuales por cliente",
        "compras_mensuales": "Compras mensuales por proveedor",
        "ventas_anio": "Ventas acumuladas año calendario",
        "ventas_12": "Ventas últimos 12 meses",
        "compras_anio": "Compras acumuladas año calendario",
        "compras_12": "Compras últimos 12 meses",
        "proveedores_anual": "Compras por proveedor y mes",
        "recategorizacion": "Recategorización de monotributo",
        "significativos": "Comprobantes significativos",
        "usd": "Comprobantes en USD",
        "iibb": "Ingresos Brutos mensual",
        "alertas": "Alertas fiscales",
        "ranking_clientes": "Ranking de clientes por ventas",
        "ranking_proveedores": "Ranking de proveedores por compras",
        "documentacion": "Documentación pendiente",
        "tareas": "Tareas pendientes",
        "honorarios": "Honorarios pendientes",
        "mercado_pago": "Reporte Mercado Pago",
        "mercado_libre": "Reporte Mercado Libre",
    }

    def __init__(self, vouchers: VoucherService) -> None:
        self.vouchers = vouchers
        self.database = vouchers.database

    def export_table_excel(
        self, destination: Path, title: str, rows: list[dict], filter_text: str = ""
    ) -> Path:
        destination.parent.mkdir(parents=True, exist_ok=True)
        dataframe = self._prepare_excel_dates(pd.DataFrame(rows or [{"Estado": "Sin información"}]))
        with pd.ExcelWriter(destination, engine="openpyxl") as writer:
            dataframe.to_excel(writer, index=False, sheet_name="Reporte", startrow=4)
        workbook = load_workbook(destination)
        sheet = workbook.active
        sheet["A1"] = title
        sheet["A2"] = "Fecha de exportación"; sheet["B2"] = date.today(); sheet["B2"].number_format = "dd/mm/yyyy"
        sheet["A3"] = "Filtro aplicado"; sheet["B3"] = filter_text or "Todos"
        sheet["D2"] = "Cantidad de registros"; sheet["E2"] = len(rows)
        sheet.freeze_panes = "A6"; sheet.auto_filter.ref = f"A5:{sheet.cell(sheet.max_row, sheet.max_column).coordinate}"
        for cell in sheet[5]:
            cell.fill = PatternFill("solid", fgColor="1F4E78"); cell.font = Font(color="FFFFFF", bold=True); cell.alignment = Alignment(horizontal="center")
        money_terms = ("importe", "saldo", "total", "comision", "retencion", "percepcion", "impuesto", "precio", "descuento", "envio", "venta", "compra", "cobranza", "pago", "transferencia", "interes", "anulacion", "nota")
        for column_index, cell in enumerate(sheet[5], 1):
            name = str(cell.value or "").casefold()
            for row_index in range(6, sheet.max_row + 1):
                target = sheet.cell(row_index, column_index)
                if "periodo" in name:
                    target.number_format = "mm/yyyy"; target.alignment = Alignment(horizontal="center")
                elif "fecha" in name:
                    target.number_format = "dd/mm/yyyy"; target.alignment = Alignment(horizontal="center")
                elif any(term in name for term in money_terms) and isinstance(target.value, (int, float)):
                    target.number_format = '#,##0.00'; target.alignment = Alignment(horizontal="right")
                else:
                    target.alignment = Alignment(horizontal="left")
        for column in sheet.columns:
            letter = column[0].column_letter
            sheet.column_dimensions[letter].width = min(max(len(str(cell.value or "")) for cell in column) + 2, 42)
        workbook.save(destination)
        return destination

    def export_table_pdf(
        self, destination: Path, title: str, rows: list[dict], filter_text: str = ""
    ) -> Path:
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.lib.units import mm
            from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
        except ImportError as error:
            raise RuntimeError("Instalá las dependencias con: python -m pip install -r requirements.txt") from error
        destination.parent.mkdir(parents=True, exist_ok=True)
        styles = getSampleStyleSheet()
        rows = rows or [{"Estado": "Sin información"}]
        headers = list(rows[0])
        if len(headers) > 12:
            preferred = (
                "cliente", "nombre_razon_social", "cuit_cuil", "fecha", "periodo",
                "descripcion", "tipo_movimiento", "tipo_operacion", "impuesto",
                "organismo", "contraparte", "producto", "importe_bruto",
                "importe_neto", "saldo", "estado", "responsable", "observaciones",
            )
            selected = [key for key in preferred if key in headers]
            headers = (selected + [key for key in headers if key not in selected])[:12]
        money_terms = ("importe", "saldo", "total", "comision", "retencion", "percepcion", "impuesto", "precio", "descuento", "envio", "venta", "compra", "cobranza", "pago", "transferencia", "interes", "anulacion", "nota")

        def formatted(key, value):
            normalized = str(key).casefold()
            if value is None: return ""
            if "periodo" in normalized: return display_period(str(value))
            if "fecha" in normalized: return display_date(str(value))
            if any(term in normalized for term in money_terms) and isinstance(value, (int, float)): return number_ar(value)
            return str(value)

        data = [[Paragraph(str(header).replace("_", " ").title(), styles["BodyText"]) for header in headers]]
        data.extend([[Paragraph(formatted(header, row.get(header)), styles["BodyText"]) for header in headers] for row in rows])
        available = 277 * mm
        widths = [available / max(len(headers), 1)] * len(headers)
        table = Table(data, colWidths=widths, repeatRows=1)
        table.setStyle(TableStyle([("BACKGROUND", (0,0), (-1,0), colors.HexColor("#1F4E78")), ("TEXTCOLOR", (0,0), (-1,0), colors.white), ("GRID", (0,0), (-1,-1), .25, colors.HexColor("#C8D0D8")), ("VALIGN", (0,0), (-1,-1), "TOP"), ("FONTSIZE", (0,0), (-1,-1), 6.5), ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, colors.HexColor("#F4F7FA")])]))
        story = [Paragraph(title, styles["Title"]), Paragraph(f"Fecha: {date.today().strftime('%d/%m/%Y')}", styles["Normal"]), Paragraph(f"Filtro: {filter_text or 'Todos'}", styles["Normal"]), Paragraph(f"Registros: {len(rows)}", styles["Normal"]), Spacer(1, 5*mm), table]
        SimpleDocTemplate(str(destination), pagesize=landscape(A4), leftMargin=10*mm, rightMargin=10*mm, topMargin=12*mm, bottomMargin=12*mm).build(story)
        return destination

    @staticmethod
    def _validate_date_range(
        date_from: str | date | None,
        date_to: str | date | None,
    ) -> tuple[date | None, date | None]:
        def parse(value: str | date | None) -> date | None:
            if not value:
                return None
            if isinstance(value, date):
                return value
            text = str(value).strip()
            for fmt in ("%Y-%m-%d", "%d/%m/%Y"):
                try:
                    return datetime.strptime(text, fmt).date()
                except ValueError:
                    continue
            raise ValueError("Las fechas del reporte deben tener formato DD/MM/AAAA.")

        start, end = parse(date_from), parse(date_to)
        if start and end and start > end:
            raise ValueError("La fecha Desde no puede ser posterior a la fecha Hasta.")
        return start, end

    @staticmethod
    def _filter_dataframe_by_range(
        dataframe: pd.DataFrame,
        start: date | None,
        end: date | None,
    ) -> pd.DataFrame:
        if dataframe.empty or (not start and not end):
            return dataframe
        candidates = (
            "periodo", "periodo_fiscal", "periodo_hasta", "fecha",
            "fecha_creacion", "fecha_vencimiento", "fecha_emision",
            "fecha_solicitud", "creado_en",
        )
        column = next((name for name in candidates if name in dataframe.columns), None)
        if not column:
            return dataframe
        values = dataframe[column].astype(str).str.strip()
        if column.startswith("periodo"):
            values = values.str.slice(0, 7) + "-01"
        parsed = pd.to_datetime(values, errors="coerce")
        mask = parsed.notna()
        if start:
            mask &= parsed.dt.date >= start
        if end:
            mask &= parsed.dt.date <= end
        return dataframe.loc[mask].reset_index(drop=True)

    @staticmethod
    def _prepare_excel_dates(dataframe: pd.DataFrame) -> pd.DataFrame:
        dataframe = dataframe.copy()
        for column in dataframe.columns:
            normalized = str(column).casefold().translate(
                str.maketrans("áéíóúñ", "aeioun")
            )
            is_period = "periodo" in normalized or normalized == "mes y ano"
            is_date = normalized.startswith("fecha") or normalized in {
                "creado_en", "actualizado_en"
            }
            if not (is_period or is_date):
                continue

            def convert(value):
                if value is None or value == "" or pd.isna(value):
                    return None
                if isinstance(value, (date, datetime, pd.Timestamp)):
                    return value.to_pydatetime() if isinstance(value, pd.Timestamp) else value
                text = str(value).strip()
                if is_period and re.fullmatch(r"\d{4}-\d{2}", text):
                    return datetime.strptime(text, "%Y-%m").date()
                for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%d/%m/%Y"):
                    try:
                        return datetime.strptime(text, fmt)
                    except ValueError:
                        continue
                return value

            dataframe[column] = dataframe[column].map(convert)
        return dataframe

    def export_named(
        self,
        report: str,
        destination: Path,
        client_id: int | None = None,
        date_from: str | date | None = None,
        date_to: str | date | None = None,
        platform_filter: str = "",
    ) -> Path:
        if report not in self.REPORTS:
            raise ValueError("El reporte seleccionado no existe.")
        if report in ("mercado_pago", "mercado_libre"):
            if not client_id:
                raise ValueError("Seleccioná un cliente para generar este reporte.")
            return self.export_platform_report(destination, report, client_id, date_from, date_to, platform_filter)
        if report == "ultimos_12_meses":
            if not client_id:
                raise ValueError("Seleccioná un cliente para generar este reporte.")
            return self.export_last_twelve_months(
                destination, client_id, date_from=date_from, date_to=date_to
            )
        if report == "proveedores_anual":
            if not client_id:
                raise ValueError("Seleccioná un cliente para generar el reporte de proveedores.")
            start, end = self._validate_date_range(date_from, date_to)
            years = {value.year for value in (start, end) if value}
            if len(years) > 1:
                raise ValueError("El reporte de proveedores debe corresponder a un único año.")
            year = next(iter(years), date.today().year)
            return self.export_supplier_matrix(destination, client_id, year)

        condition = " AND cliente_id = ?" if client_id else ""
        params = (client_id,) if client_id else ()
        start, end = self._validate_date_range(date_from, date_to)
        if not start and not end and report in ("ventas_anio", "compras_anio"):
            today = date.today()
            start, end = date(today.year, 1, 1), date(today.year, 12, 31)
        elif not start and not end and report in ("ventas_12", "compras_12"):
            today = date.today()
            start = date(today.year - 1, today.month, 1)
            end = today
        date_condition = ""
        date_params: list[str] = []
        if start:
            date_condition += " AND date(fecha) >= date(?)"
            date_params.append(start.isoformat())
        if end:
            date_condition += " AND date(fecha) <= date(?)"
            date_params.append(end.isoformat())
        voucher_condition = condition + date_condition
        voucher_params = (*params, *date_params)
        queries = {
            "ventas_mensuales": (
                "SELECT periodo_fiscal AS periodo, contraparte_nombre, "
                "contraparte_documento, COUNT(*) cantidad, "
                "SUM(importe_neto_fiscal) total FROM comprobantes_ventas "
                "WHERE 1=1" + voucher_condition + " GROUP BY periodo_fiscal, "
                "contraparte_nombre, contraparte_documento "
                "ORDER BY periodo_fiscal DESC, total DESC"
            ),
            "compras_mensuales": (
                "SELECT periodo_fiscal AS periodo, contraparte_nombre, "
                "contraparte_documento, COUNT(*) cantidad, "
                "SUM(importe_neto_fiscal) total FROM comprobantes_compras "
                "WHERE 1=1" + voucher_condition + " GROUP BY periodo_fiscal, "
                "contraparte_nombre, contraparte_documento "
                "ORDER BY periodo_fiscal DESC, total DESC"
            ),
            "ventas_anio": "SELECT * FROM comprobantes_ventas WHERE 1=1" + voucher_condition + " ORDER BY fecha",
            "ventas_12": "SELECT * FROM comprobantes_ventas WHERE 1=1" + voucher_condition + " ORDER BY fecha",
            "compras_anio": "SELECT * FROM comprobantes_compras WHERE 1=1" + voucher_condition + " ORDER BY fecha",
            "compras_12": "SELECT * FROM comprobantes_compras WHERE 1=1" + voucher_condition + " ORDER BY fecha",
            "recategorizacion": (
                "SELECT * FROM recategorizaciones_monotributo WHERE 1=1"
                + condition + " ORDER BY creado_en DESC"
            ),
            "usd": (
                "SELECT 'Venta' operacion, * FROM comprobantes_ventas "
                "WHERE moneda='USD'" + voucher_condition + " UNION ALL "
                "SELECT 'Compra' operacion, * FROM comprobantes_compras "
                "WHERE moneda='USD'" + voucher_condition
            ),
            "iibb": "SELECT * FROM iibb_monotributo WHERE 1=1" + condition + " ORDER BY periodo DESC",
            "alertas": "SELECT * FROM alertas_fiscales WHERE 1=1" + condition + " ORDER BY fecha_creacion DESC",
            "documentacion": "SELECT * FROM documentacion WHERE estado NOT IN ('Aprobada','aprobada')" + condition + " ORDER BY id DESC",
            "tareas": "SELECT * FROM tareas WHERE estado NOT IN ('finalizado','archivado','cobrado')" + condition + " ORDER BY fecha_vencimiento",
            "honorarios": "SELECT * FROM honorarios WHERE estado NOT IN ('cobrado total')" + condition + " ORDER BY id DESC",
            "ranking_clientes": (
                "SELECT contraparte_nombre, contraparte_documento, COUNT(*) cantidad, "
                "SUM(importe_neto_fiscal) total FROM comprobantes_ventas WHERE 1=1"
                + voucher_condition + " GROUP BY contraparte_nombre, contraparte_documento ORDER BY total DESC"
            ),
            "ranking_proveedores": (
                "SELECT contraparte_nombre, contraparte_documento, COUNT(*) cantidad, "
                "SUM(importe_neto_fiscal) total FROM comprobantes_compras WHERE 1=1"
                + voucher_condition + " GROUP BY contraparte_nombre, contraparte_documento ORDER BY total DESC"
            ),
        }
        voucher_reports = {
            "ventas_mensuales", "compras_mensuales", "ventas_anio", "ventas_12",
            "compras_anio", "compras_12", "ranking_clientes", "ranking_proveedores",
            "significativos", "usd",
        }
        if report == "significativos":
            threshold = (
                self.vouchers.config.get_client_float(
                    client_id, "monto_comprobante_significativo", 500_000
                )
                if client_id
                else self.vouchers.config.get_float(
                    "monto_comprobante_significativo", 500_000
                )
            )
            sql = (
                "SELECT 'Venta' operacion, * FROM comprobantes_ventas "
                "WHERE ABS(importe_pesos)>=?" + voucher_condition + " UNION ALL "
                "SELECT 'Compra' operacion, * FROM comprobantes_compras "
                "WHERE ABS(importe_pesos)>=?" + voucher_condition
            )
            query_params = (threshold, *voucher_params, threshold, *voucher_params)
        else:
            sql = queries[report]
            if report == "usd":
                query_params = (*voucher_params, *voucher_params)
            elif report in voucher_reports:
                query_params = voucher_params
            else:
                query_params = params
        rows = [dict(row) for row in self.database.query(sql, query_params)]
        if not rows:
            raise ValueError("No hay datos para el reporte seleccionado.")
        dataframe = pd.DataFrame(rows)
        if report not in voucher_reports and report != "usd":
            dataframe = self._filter_dataframe_by_range(dataframe, start, end)
        if dataframe.empty:
            raise ValueError("No hay datos dentro del rango de fechas seleccionado.")
        dataframe = self._prepare_excel_dates(dataframe)
        return self._write_dataframe(destination, dataframe, "Reporte")

    def platform_report_data(
        self, report: str, client_id: int,
        date_from: str | date | None = None, date_to: str | date | None = None,
        platform_filter: str = "",
    ) -> dict[str, list[dict]]:
        start, end = self._validate_date_range(date_from, date_to)
        platform = PlatformService(self.database)
        if report == "mercado_pago":
            detail = platform.list_mp(client_id)
            if start: detail = [row for row in detail if row["fecha"] >= start.isoformat()]
            if end: detail = [row for row in detail if row["fecha"] <= end.isoformat()]
            if platform_filter:
                term = platform_filter.casefold()
                detail = [row for row in detail if term in " ".join((str(row.get("tipo_movimiento","")),str(row.get("ingreso_egreso","")),str(row.get("contraparte","")),str(row.get("estado","")))).casefold()]
            return {
                "Movimientos": detail,
                "Resumen Mensual": platform.mp_summary(client_id),
                "Ranking Ingresos": platform.mp_ranking(client_id, "Ingreso"),
                "Ranking Egresos": platform.mp_ranking(client_id, "Egreso"),
                "Significativos": [row for row in detail if abs(float(row["importe_neto"] or 0)) >= 1000],
                "A revisar": [row for row in detail if row["tipo_movimiento"] == "A revisar"],
            }
        detail = platform.list_ml(client_id)
        if start: detail = [row for row in detail if row["fecha"] >= start.isoformat()]
        if end: detail = [row for row in detail if row["fecha"] <= end.isoformat()]
        if platform_filter:
            term = platform_filter.casefold()
            detail = [row for row in detail if term in " ".join((str(row.get("tipo_operacion","")),str(row.get("contraparte","")),str(row.get("producto","")),str(row.get("estado","")))).casefold()]
        products = {}
        counterparts = {}
        for row in detail:
            product = row["producto"] or "Sin identificar"
            products.setdefault(product, {"producto": product, "cantidad": 0.0, "importe_total": 0.0})
            products[product]["cantidad"] += float(row["cantidad"] or 0); products[product]["importe_total"] += float(row["importe_neto"] or 0)
            counterpart = row["contraparte"] or "Sin identificar"
            counterparts.setdefault(counterpart, {"contraparte": counterpart, "operaciones": 0, "importe_total": 0.0})
            counterparts[counterpart]["operaciones"] += 1; counterparts[counterpart]["importe_total"] += float(row["importe_neto"] or 0)
        return {
            "Operaciones": detail,
            "Resumen Mensual": platform.ml_summary(client_id),
            "Productos": sorted(products.values(), key=lambda item: item["importe_total"], reverse=True),
            "Compradores": sorted(counterparts.values(), key=lambda item: item["importe_total"], reverse=True),
            "Significativos": [row for row in detail if abs(float(row["importe_neto"] or 0)) >= 500000],
            "A revisar": [row for row in detail if not row["id_operacion"] and not row["id_venta"] and not row["numero_comprobante"]],
        }

    def export_platform_report(
        self, destination: Path, report: str, client_id: int,
        date_from: str | date | None = None, date_to: str | date | None = None,
        platform_filter: str = "",
    ) -> Path:
        sheets = self.platform_report_data(report, client_id, date_from, date_to, platform_filter)
        if not any(sheets.values()):
            raise ValueError("No hay datos para el reporte seleccionado.")
        destination.parent.mkdir(parents=True, exist_ok=True)
        with pd.ExcelWriter(destination, engine="openpyxl") as writer:
            for name, rows in sheets.items():
                dataframe = self._prepare_excel_dates(pd.DataFrame(rows or [{"Estado": "Sin información"}]))
                dataframe.to_excel(writer, sheet_name=name[:31], index=False, startrow=3)
        workbook = load_workbook(destination)
        client = self.database.query_one("SELECT nombre_razon_social,cuit_cuil FROM clientes WHERE id=?", (client_id,))
        for sheet in workbook.worksheets:
            sheet["A1"] = self.REPORTS[report]; sheet["B1"] = client["nombre_razon_social"] if client else ""
            sheet["A2"] = "CUIT / CUIL"; sheet["B2"] = client["cuit_cuil"] if client else ""
            sheet["D1"] = "Fecha de emisión"; sheet["E1"] = date.today(); sheet["E1"].number_format = "dd/mm/yyyy"
            sheet.freeze_panes = "A5"; sheet.auto_filter.ref = f"A4:{sheet.cell(sheet.max_row, sheet.max_column).coordinate}"
            for cell in sheet[4]: cell.fill = PatternFill("solid", fgColor="1F4E78"); cell.font = Font(color="FFFFFF", bold=True)
            for column_index, header in enumerate(sheet[4], 1):
                name = str(header.value or "").casefold()
                for row_index in range(5, sheet.max_row + 1):
                    cell = sheet.cell(row_index, column_index)
                    if "periodo" in name: cell.number_format = "mm/yyyy"; cell.alignment = Alignment(horizontal="center")
                    elif "fecha" in name: cell.number_format = "dd/mm/yyyy"; cell.alignment = Alignment(horizontal="center")
                    elif any(term in name for term in ("importe", "saldo", "total", "comision", "retencion", "percepcion", "precio", "descuento", "envio", "venta", "compra", "cobranza", "pago", "transferencia", "interes", "anulacion", "nota")) and isinstance(cell.value, (int,float)):
                        cell.number_format = '#,##0.00'; cell.alignment = Alignment(horizontal="right")
            for column in sheet.columns:
                sheet.column_dimensions[column[0].column_letter].width = min(max(len(str(cell.value or "")) for cell in column)+2, 38)
        workbook.save(destination)
        return destination

    def export_named_pdf(
        self, report: str, destination: Path, client_id: int | None = None,
        date_from: str | date | None = None, date_to: str | date | None = None,
        platform_filter: str = "",
    ) -> Path:
        if report in ("mercado_pago", "mercado_libre"):
            if not client_id: raise ValueError("Seleccioná un cliente.")
            sheets = self.platform_report_data(report, client_id, date_from, date_to, platform_filter)
            rows = sheets.get("Resumen Mensual", []) or sheets.get("Movimientos", []) or sheets.get("Operaciones", [])
            return self.export_table_pdf(destination, self.REPORTS[report], rows, f"Cliente ID {client_id}")
        # Para reportes existentes, se genera primero la misma consulta en un Excel
        # temporal y se presenta un resumen tabular del contenido.
        import tempfile
        with tempfile.TemporaryDirectory() as directory:
            xlsx = Path(directory) / "reporte.xlsx"
            self.export_named(report, xlsx, client_id, date_from, date_to)
            frame = pd.read_excel(xlsx)
        return self.export_table_pdf(destination, self.REPORTS[report], frame.to_dict("records"), "Rango seleccionado")

    @staticmethod
    def _last_twelve_periods(reference: date | None = None) -> list[str]:
        reference = reference or date.today()
        current = reference.year * 12 + reference.month - 1
        result = []
        for month_index in range(current - 11, current + 1):
            year, zero_based_month = divmod(month_index, 12)
            result.append(f"{year}-{zero_based_month + 1:02d}")
        return result

    def last_twelve_months(
        self,
        client_id: int,
        reference: date | None = None,
        date_from: str | date | None = None,
        date_to: str | date | None = None,
    ) -> dict:
        if not self.database.query_one(
            "SELECT id FROM clientes WHERE id = ?", (client_id,)
        ):
            raise ValueError("El cliente seleccionado no existe.")
        start, end = self._validate_date_range(date_from, date_to)
        if start or end:
            end = end or reference or date.today()
            if not start:
                end_index = end.year * 12 + end.month - 1
                start_year, start_month_zero = divmod(end_index - 11, 12)
                start = date(start_year, start_month_zero + 1, 1)
            start_index = start.year * 12 + start.month - 1
            end_index = end.year * 12 + end.month - 1
            periods = []
            for month_index in range(start_index, end_index + 1):
                year, month_zero = divmod(month_index, 12)
                periods.append(f"{year}-{month_zero + 1:02d}")
        else:
            periods = self._last_twelve_periods(reference)
        placeholders = ",".join("?" for _ in periods)

        def totals_by_period(table: str) -> dict[str, float]:
            return {
                row["periodo"]: float(row["total"] or 0)
                for row in self.database.query(
                    f"""
                    SELECT periodo_fiscal AS periodo,
                           COALESCE(SUM(importe_neto_fiscal), 0) AS total
                    FROM {table}
                    WHERE cliente_id = ? AND periodo_fiscal IN ({placeholders})
                    GROUP BY periodo_fiscal
                    """,
                    (client_id, *periods),
                )
            }

        sales = totals_by_period("comprobantes_ventas")
        purchases = totals_by_period("comprobantes_compras")
        fixed_amounts = {
            row["periodo"]: float(row["importe_fijo"] or 0)
            for row in self.database.query(
                f"""
                SELECT periodo, importe_fijo FROM iibb_monotributo
                WHERE cliente_id = ? AND periodo IN ({placeholders})
                """,
                (client_id, *periods),
            )
        }
        profile = self.database.query_one(
            "SELECT alicuota FROM ingresos_brutos_cliente WHERE cliente_id = ?",
            (client_id,),
        )
        default_rate = self.vouchers.config.get_float("alicuota_iibb_default", 0.035)
        rate = float(profile["alicuota"] or default_rate) if profile else default_rate
        rows = []
        for period in periods:
            year, month = (int(value) for value in period.split("-"))
            sales_total = sales.get(period, 0.0)
            purchases_total = purchases.get(period, 0.0)
            rows.append(
                {
                    "periodo": period,
                    "mes": f"{month:02d}/{year}",
                    "ventas": round(sales_total, 2),
                    "compras": round(purchases_total, 2),
                    "resultado": round(sales_total - purchases_total, 2),
                    "porcentaje_compras": (
                        purchases_total / sales_total if sales_total else 0.0
                    ),
                    "ingresos_brutos": round(sales_total * rate, 2),
                    "regimen_simplificado": round(fixed_amounts.get(period, 0.0), 2),
                }
            )
        total_sales = sum(row["ventas"] for row in rows)
        total_purchases = sum(row["compras"] for row in rows)
        totals = {
            "mes": "TOTAL",
            "ventas": round(total_sales, 2),
            "compras": round(total_purchases, 2),
            "resultado": round(total_sales - total_purchases, 2),
            "porcentaje_compras": (
                total_purchases / total_sales if total_sales else 0.0
            ),
            "ingresos_brutos": round(sum(row["ingresos_brutos"] for row in rows), 2),
            "regimen_simplificado": round(
                sum(row["regimen_simplificado"] for row in rows), 2
            ),
        }
        return {"rows": rows, "totals": totals, "rate": rate, "periods": periods}

    def export_last_twelve_months(
        self,
        destination: Path,
        client_id: int,
        reference: date | None = None,
        date_from: str | date | None = None,
        date_to: str | date | None = None,
    ) -> Path:
        report = self.last_twelve_months(
            client_id, reference, date_from=date_from, date_to=date_to
        )
        rows = [
            {
                "Mes y año": date(
                    int(row["periodo"][:4]), int(row["periodo"][5:7]), 1
                ),
                "Ventas": row["ventas"],
                "Compras": row["compras"],
                "Resultado": row["resultado"],
                "Compras / ventas": row["porcentaje_compras"],
                "Ingresos Brutos": row["ingresos_brutos"],
                "Régimen simplificado": row["regimen_simplificado"],
            }
            for row in report["rows"]
        ]
        totals = report["totals"]
        rows.append(
            {
                "Mes y año": "TOTAL",
                "Ventas": totals["ventas"],
                "Compras": totals["compras"],
                "Resultado": totals["resultado"],
                "Compras / ventas": totals["porcentaje_compras"],
                "Ingresos Brutos": totals["ingresos_brutos"],
                "Régimen simplificado": totals["regimen_simplificado"],
            }
        )
        dataframe = pd.DataFrame(rows)
        currency = '"$"#,##0.00;[Red]-"$"#,##0.00'
        self._write_dataframe(
            destination,
            dataframe,
            "Últimos 12 meses",
            {
                "Mes y año": "mm/yyyy",
                "Ventas": currency,
                "Compras": currency,
                "Resultado": currency,
                "Compras / ventas": "0.0%",
                "Ingresos Brutos": currency,
                "Régimen simplificado": currency,
            },
        )
        workbook = load_workbook(destination)
        sheet = workbook["Últimos 12 meses"]
        total_fill = PatternFill("solid", fgColor="D9EAF7")
        total_border = Border(top=Side(style="medium", color="1F4E78"))
        for cell in sheet[sheet.max_row]:
            cell.fill = total_fill
            cell.font = Font(bold=True, color="17324D")
            cell.border = total_border
        workbook.save(destination)
        return destination

    def export_client_file(self, destination: Path, client_id: int) -> Path:
        """Exporta la ficha existente del cliente en hojas separadas y auditables."""
        client_row = self.database.query_one(
            """
            SELECT c.*, df.regimen_principal, df.condicion_iva,
                   df.domicilio_fiscal, df.observaciones AS observaciones_fiscales
            FROM clientes c
            LEFT JOIN datos_fiscales_cliente df ON df.cliente_id = c.id
            WHERE c.id = ?
            """,
            (client_id,),
        )
        if not client_row:
            raise ValueError("El cliente seleccionado no existe.")
        client = dict(client_row)

        def records(sql: str) -> list[dict]:
            return [dict(row) for row in self.database.query(sql, (client_id,))]

        profiles = {
            "Resumen": {
                "Nombre / Razón social": client.get("nombre_razon_social", ""),
                "CUIT / CUIL": client.get("cuit_cuil", ""),
                "Tipo de persona": client.get("tipo_persona", ""),
                "Condición fiscal": client.get("regimen_principal", ""),
                "Estado": client.get("estado", ""),
                "Actividad / rubro": client.get("rubro") or client.get("actividad", ""),
                "Fecha de alta": client.get("fecha_alta_estudio", ""),
                "Observaciones": client.get("observaciones", ""),
                "Fecha de exportación": date.today(),
            },
            "Datos Cliente": client,
        }
        sheets: dict[str, pd.DataFrame] = {
            name: pd.DataFrame(
                [{"Campo": key, "Valor": value} for key, value in values.items()]
            )
            for name, values in profiles.items()
        }
        queries = {
            "Monotributo": "SELECT * FROM monotributo_cliente WHERE cliente_id=?",
            "IIBB": "SELECT * FROM ingresos_brutos_cliente WHERE cliente_id=?",
            "IIBB Mensual": "SELECT * FROM iibb_monotributo WHERE cliente_id=? ORDER BY periodo DESC",
            "Ventas": "SELECT * FROM comprobantes_ventas WHERE cliente_id=? ORDER BY fecha DESC",
            "Compras": "SELECT * FROM comprobantes_compras WHERE cliente_id=? ORDER BY fecha DESC",
            "Alertas": "SELECT * FROM alertas_fiscales WHERE cliente_id=? ORDER BY fecha_creacion DESC",
            "Documentación": "SELECT * FROM documentacion WHERE cliente_id=? ORDER BY id DESC",
            "Tareas": "SELECT * FROM tareas WHERE cliente_id=? ORDER BY id DESC",
            "Vencimientos": "SELECT * FROM vencimientos WHERE cliente_id=? ORDER BY fecha_vencimiento",
            "Honorarios": "SELECT * FROM honorarios WHERE cliente_id=? ORDER BY id DESC",
            "Historial importaciones": "SELECT * FROM importaciones_archivos WHERE cliente_id=? ORDER BY fecha_importacion DESC",
        }
        for sheet_name, sql in queries.items():
            rows = records(sql)
            sheets[sheet_name] = (
                pd.DataFrame(rows)
                if rows
                else pd.DataFrame([{"Estado": "Sin información cargada"}])
            )

        destination.parent.mkdir(parents=True, exist_ok=True)
        with pd.ExcelWriter(destination, engine="openpyxl") as writer:
            for sheet_name, dataframe in sheets.items():
                prepared = self._prepare_excel_dates(dataframe)
                prepared.to_excel(writer, sheet_name=sheet_name[:31], index=False)
        workbook = load_workbook(destination)
        header_fill = PatternFill("solid", fgColor="1F4E78")
        for sheet in workbook.worksheets:
            sheet.sheet_view.showGridLines = False
            sheet.freeze_panes = "A2"
            sheet.auto_filter.ref = sheet.dimensions
            for cell in sheet[1]:
                cell.fill = header_fill
                cell.font = Font(color="FFFFFF", bold=True)
            for column in sheet.columns:
                width = min(max(len(str(cell.value or "")) for cell in column) + 2, 45)
                sheet.column_dimensions[column[0].column_letter].width = max(width, 12)
            headers = {cell.value: cell.column for cell in sheet[1]}
            for header, column_index in headers.items():
                normalized = str(header).casefold().translate(
                    str.maketrans("áéíóúñ", "aeioun")
                )
                if "periodo" in normalized:
                    number_format = "mm/yyyy"
                elif normalized.startswith("fecha") or normalized.endswith("_en"):
                    number_format = "dd/mm/yyyy"
                else:
                    continue
                for row_index in range(2, sheet.max_row + 1):
                    sheet.cell(row=row_index, column=column_index).number_format = number_format
        workbook.save(destination)
        return destination

    def supplier_matrix(self, client_id: int, year: int) -> pd.DataFrame:
        rows = self.database.query(
            """
            SELECT contraparte_nombre AS proveedor,
                   CAST(substr(periodo_fiscal, 6, 2) AS INTEGER) AS mes,
                   SUM(importe_neto_fiscal) AS importe
            FROM comprobantes_compras
            WHERE cliente_id = ? AND periodo_fiscal LIKE ?
            GROUP BY contraparte_nombre, substr(periodo_fiscal, 6, 2)
            ORDER BY contraparte_nombre COLLATE NOCASE
            """,
            (client_id, f"{int(year):04d}-%"),
        )
        if not rows:
            raise ValueError("No hay compras de proveedores para el año seleccionado.")
        months = (
            "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
            "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre",
        )
        providers: dict[str, list[float]] = {}
        for row in rows:
            provider = str(row["proveedor"] or "SIN PROVEEDOR")
            providers.setdefault(provider, [0.0] * 12)
            month = int(row["mes"] or 0)
            if 1 <= month <= 12:
                providers[provider][month - 1] += float(row["importe"] or 0)
        data = []
        for provider, values in providers.items():
            item = {"Proveedor": provider}
            item.update({month: round(values[index], 2) for index, month in enumerate(months)})
            item["Total"] = round(sum(values), 2)
            data.append(item)
        totals = {"Proveedor": "TOTAL"}
        for month in months:
            totals[month] = round(sum(item[month] for item in data), 2)
        totals["Total"] = round(sum(item["Total"] for item in data), 2)
        data.append(totals)
        return pd.DataFrame(data, columns=("Proveedor", *months, "Total"))

    def export_supplier_matrix(
        self, destination: Path, client_id: int, year: int
    ) -> Path:
        dataframe = self.supplier_matrix(client_id, year)
        currency = '"$"#,##0.00;[Red]-"$"#,##0.00'
        formats = {
            column: currency for column in dataframe.columns if column != "Proveedor"
        }
        self._write_dataframe(
            destination, dataframe, f"Proveedores {int(year)}", formats
        )
        workbook = load_workbook(destination)
        sheet = workbook[f"Proveedores {int(year)}"]
        for cell in sheet[sheet.max_row]:
            cell.fill = PatternFill("solid", fgColor="D9EAF7")
            cell.font = Font(bold=True, color="17324D")
            cell.border = Border(top=Side(style="medium", color="1F4E78"))
        sheet.freeze_panes = "B2"
        workbook.save(destination)
        return destination

    def export_vouchers(
        self,
        kind: str,
        destination: Path,
        client_id: int | None = None,
        period: str = "",
    ) -> Path:
        rows = self.vouchers.list(kind, client_id=client_id, period=period)
        if not rows:
            raise ValueError("No hay comprobantes para exportar con los filtros actuales.")
        columns = [
            "cliente_nombre", "fecha", "periodo_fiscal", "tipo_comprobante",
            "punto_venta", "numero_comprobante", "contraparte_nombre",
            "contraparte_documento", "moneda", "tipo_cambio", "importe_original",
            "importe_pesos", "importe_neto_fiscal", "estado", "origen", "observaciones",
        ]
        labels = {
            "cliente_nombre": "Cliente del estudio",
            "fecha": "Fecha",
            "periodo_fiscal": "Período",
            "tipo_comprobante": "Tipo",
            "punto_venta": "Punto de venta",
            "numero_comprobante": "Número",
            "contraparte_nombre": "Cliente / proveedor",
            "contraparte_documento": "CUIT / DNI",
            "moneda": "Moneda",
            "tipo_cambio": "Tipo de cambio",
            "importe_original": "Importe original",
            "importe_pesos": "Importe en pesos",
            "importe_neto_fiscal": "Importe neto fiscal",
            "estado": "Estado",
            "origen": "Origen",
            "observaciones": "Observaciones",
        }
        dataframe = pd.DataFrame(rows)[columns].rename(columns=labels)
        currency = '"$"#,##0.00;[Red]-"$"#,##0.00'
        return self._write_dataframe(
            destination,
            dataframe,
            kind.capitalize(),
            {
                "Importe original": currency,
                "Importe en pesos": currency,
                "Importe neto fiscal": currency,
            },
        )

    def export_accounting_section(
        self,
        kind: str,
        section: str,
        destination: Path,
        client_id: int,
    ) -> Path:
        configurations = {
            "resumen_mensual": {
                "sheet": "Resumen mensual",
                "rows": lambda: self.vouchers.monthly_summary(kind, client_id),
                "columns": ("periodo", "facturas", "notas_credito", "notas_debito", "anulados", "total_neto", "cantidad"),
                "labels": ("Período", "Facturas", "Notas de crédito", "Notas de débito", "Comprobantes anulados", "Neto", "Cantidad"),
                "formats": {"Facturas": "currency", "Notas de crédito": "currency", "Notas de débito": "currency", "Neto": "currency", "Comprobantes anulados": "#,##0", "Cantidad": "#,##0"},
            },
            "significativos": {
                "sheet": "Significativos",
                "rows": lambda: self.vouchers.noteworthy(kind, client_id),
                "columns": ("fecha", "tipo_comprobante", "contraparte_nombre", "contraparte_documento", "moneda", "importe_pesos", "motivo_alerta"),
                "labels": ("Fecha", "Tipo de comprobante", "Cliente / proveedor", "CUIT / DNI", "Moneda", "Importe en pesos", "Motivo"),
                "formats": {"Importe en pesos": "currency"},
            },
            "moneda_extranjera": {
                "sheet": "Moneda extranjera",
                "rows": lambda: self.vouchers.noteworthy(kind, client_id, True),
                "columns": ("fecha", "tipo_comprobante", "contraparte_nombre", "moneda", "importe_original", "tipo_cambio", "importe_pesos", "estado"),
                "labels": ("Fecha", "Tipo de comprobante", "Cliente / proveedor", "Moneda", "Importe original", "Tipo de cambio", "Importe en pesos", "Estado"),
                "formats": {"Importe original": "#,##0.00;[Red]-#,##0.00", "Tipo de cambio": "#,##0.0000", "Importe en pesos": "currency"},
            },
            "ranking": {
                "sheet": "Ranking",
                "rows": lambda: self.vouchers.ranking(kind, client_id),
                "columns": ("puesto", "contraparte_nombre", "contraparte_documento", "total", "cantidad", "porcentaje"),
                "labels": ("Puesto", "Cliente / proveedor", "CUIT / DNI", "Total", "Cantidad", "Participación"),
                "formats": {"Puesto": "#,##0", "Total": "currency", "Cantidad": "#,##0", "Participación": "0.0%"},
            },
        }
        if section not in configurations:
            raise ValueError("La sección contable seleccionada no existe.")
        if not client_id:
            raise ValueError("Seleccioná un cliente para exportar esta sección.")
        configuration = configurations[section]
        rows = configuration["rows"]()
        if not rows:
            raise ValueError("No hay datos para exportar en esta sección.")
        dataframe = pd.DataFrame(rows)[list(configuration["columns"])]
        dataframe.columns = list(configuration["labels"])
        currency = '"$"#,##0.00;[Red]-"$"#,##0.00'
        formats = {
            column: currency if value == "currency" else value
            for column, value in configuration["formats"].items()
        }
        return self._write_dataframe(
            destination, dataframe, str(configuration["sheet"]), formats
        )

    @staticmethod
    def _write_dataframe(
        destination: Path,
        dataframe: pd.DataFrame,
        sheet_name: str,
        number_formats: dict[str, str] | None = None,
    ) -> Path:
        dataframe = ReportService._prepare_excel_dates(dataframe)
        formats = dict(number_formats or {})
        for column in dataframe.columns:
            normalized = str(column).casefold().translate(
                str.maketrans("áéíóúñ", "aeioun")
            )
            if "periodo" in normalized or normalized == "mes y ano":
                formats.setdefault(column, "mm/yyyy")
            elif normalized.startswith("fecha") or normalized in {
                "creado_en", "actualizado_en"
            }:
                formats.setdefault(column, "dd/mm/yyyy")
        destination.parent.mkdir(parents=True, exist_ok=True)
        with pd.ExcelWriter(destination, engine="openpyxl") as writer:
            dataframe.to_excel(writer, sheet_name=sheet_name, index=False)
        workbook = load_workbook(destination)
        sheet = workbook[sheet_name]
        sheet.sheet_view.showGridLines = False
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        sheet.row_dimensions[1].height = 24
        header_fill = PatternFill("solid", fgColor="1F4E78")
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = Font(color="FFFFFF", bold=True)
        for column in sheet.columns:
            width = min(max(len(str(cell.value or "")) for cell in column) + 2, 38)
            sheet.column_dimensions[column[0].column_letter].width = max(width, 11)
        for column_name, number_format in formats.items():
            column_index = dataframe.columns.get_loc(column_name) + 1
            for row_index in range(2, sheet.max_row + 1):
                sheet.cell(row=row_index, column=column_index).number_format = number_format
        workbook.save(destination)
        return destination
