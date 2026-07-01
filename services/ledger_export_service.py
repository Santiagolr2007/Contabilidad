from __future__ import annotations

import re
import tempfile
import zipfile
from datetime import date, datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

from database import Database
from .ledger_service import LedgerService
from .responsible_profile import RESPONSIBLE_PROFILE_SECTIONS


class LedgerExportService:
    SECTION_ORDER = (
        "resumen", "datos_cliente", "responsable_inscripto",
        *LedgerService.VISIBLE_SECTIONS,
    )

    def __init__(self, database: Database, ledger: LedgerService) -> None:
        self.database = database
        self.ledger = ledger

    @staticmethod
    def safe_name(value: str) -> str:
        return re.sub(r'[<>:"/\\|?*;]+', "_", value).strip(" ._") or "Cliente"

    def section_rows(self, client_id: int, section: str) -> list[dict]:
        if section not in self.SECTION_ORDER:
            raise ValueError("La sección seleccionada no forma parte del legajo activo.")
        if section == "resumen":
            summary = self.ledger.summary(client_id)
            client = summary.pop("client")
            area_states = summary.pop("estados_area", {})
            for hidden in ("responsable_interno", "ultima_actualizacion"):
                summary.pop(hidden, None)
            values = {
                "Cliente": client["nombre_razon_social"],
                "Legajo": client.get("legajo", ""),
                "CUIT": client["cuit_cuil"],
                "Tipo": summary.get("tipo_cliente", client["tipo_persona"]),
                "Condición fiscal": str(client["regimen_principal"]).replace("_", " ").title(),
                "Estado cliente": summary.get("estado_cliente", client["estado"]),
                **{key.replace("_", " ").title(): value for key, value in summary.items()},
                **{f"Estado {key}": value for key, value in area_states.items()},
                "Fecha exportación": date.today(),
            }
            return [{"Campo": key, "Valor": value} for key, value in values.items()]
        if section == "datos_cliente":
            row = self.database.query_one(
                "SELECT * FROM clientes WHERE id=?", (client_id,)
            )
            hidden = {"id", "creado_en", "actualizado_en"}
            values = dict(row) if row else {}
            ordered = (
                ("fecha_alta_estudio", "Fecha de alta en el estudio"), ("legajo", "Legajo"),
                ("nombre_razon_social", "Cliente"), ("cuit_cuil", "CUIT"), ("dni", "DNI"),
                ("fecha_nacimiento", "Fecha de nacimiento"), ("nacionalidad", "Nacionalidad"),
                ("estado_civil", "Estado civil"), ("tipo_persona_detalle", "Tipo de persona"),
                ("instagram", "Instagram"), ("telefono", "Teléfono"), ("email", "Mail"),
                ("domicilio", "Domicilio"), ("codigo_actividad", "Código de actividad"),
                ("actividad", "Actividad"), ("estado_detalle", "Estado"), ("rubro", "Rubro"),
                ("observaciones", "Observaciones"),
            )
            result = [{"Campo": label, "Valor": values.get(key, "")} for key, label in ordered]
            records = self.ledger.list_records(client_id, "datos_complementarios")
            if records:
                fields = self.ledger.SECTIONS["datos_complementarios"][1]
                labels = {key: label for key, label, _options in fields}
                result.extend(
                    {"Campo": labels[key], "Valor": value}
                    for key, value in records[0]["datos"].items()
                    if key in labels and value not in (None, "")
                )
            return result
        if section == "responsable_inscripto":
            values = {
                str(row["campo"]): row["valor"]
                for row in self.database.query(
                    """SELECT campo,valor FROM cliente_legajo_campos
                       WHERE cliente_id=? AND seccion='responsable_inscripto'""",
                    (client_id,),
                )
            }
            rows = []
            for group, fields in RESPONSIBLE_PROFILE_SECTIONS:
                for key, label, _kind in fields:
                    rows.append({"Sección": group, "Campo": label, "Valor": values.get(key, "")})
            return rows
        rows = self.ledger.list_records(client_id, section)
        flattened = []
        for row in rows:
            base = {"estado": row["estado"]}
            base.update(row["datos"])
            for hidden in ("responsable", "responsable_interno", "ultima_actualizacion", "actualizado_en"):
                base.pop(hidden, None)
            flattened.append(base)
        if flattened and section == "pagos":
            flattened.append({
                "estado": "TOTALES",
                "importe_facturado": sum(float(row.get("importe_facturado") or 0) for row in flattened),
                "importe_cobrado": sum(float(row.get("importe_cobrado") or 0) for row in flattened),
                "saldo_pendiente": sum(float(row.get("saldo_pendiente") or 0) for row in flattened),
            })
        return flattened

    def section_title(self, section: str) -> str:
        if section == "resumen": return "Resumen"
        if section == "datos_cliente": return "Datos Cliente"
        if section == "responsable_inscripto": return "Responsable Inscripto"
        return self.ledger.SECTIONS[section][0]

    def export_excel(
        self, destination: Path, client_id: int, sections: list[str] | None = None
    ) -> Path:
        sections = sections or list(self.SECTION_ORDER)
        destination.parent.mkdir(parents=True, exist_ok=True)
        summary = self.ledger.summary(client_id)
        client = summary["client"]
        with pd.ExcelWriter(destination, engine="openpyxl") as writer:
            for section in sections:
                rows = self.section_rows(client_id, section)
                dataframe = pd.DataFrame(rows or [{"Estado": "Sin información cargada"}])
                for column in dataframe.columns:
                    name = str(column).casefold()
                    if "periodo" in name:
                        dataframe[column] = dataframe[column].map(lambda value: datetime.strptime(value, "%Y-%m").date() if isinstance(value, str) and re.fullmatch(r"\d{4}-\d{2}", value) else value)
                    elif name.startswith("fecha") or name in ("creado_en", "actualizado_en"):
                        def convert_date(value):
                            if value is None or value == "": return None
                            if isinstance(value, (date, datetime, pd.Timestamp)): return value
                            text = str(value).strip()
                            for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%d/%m/%Y"):
                                try: return datetime.strptime(text, fmt)
                                except ValueError: continue
                            return value
                        dataframe[column] = dataframe[column].map(convert_date)
                dataframe.to_excel(
                    writer, sheet_name=self.section_title(section)[:31], index=False,
                    startrow=5,
                )
        workbook = load_workbook(destination)
        for sheet in workbook.worksheets:
            sheet.sheet_view.showGridLines = False
            sheet["A1"] = "Cliente"; sheet["B1"] = client["nombre_razon_social"]
            sheet["A2"] = "Legajo"; sheet["B2"] = client.get("legajo", "")
            sheet["A3"] = "CUIT"; sheet["B3"] = client["cuit_cuil"]
            sheet["A4"] = "Fecha de exportación"; sheet["B4"] = date.today()
            sheet["B4"].number_format = "dd/mm/yyyy"
            for row in range(1, 5):
                sheet.cell(row, 1).font = Font(bold=True, color="1F4E78")
            sheet.freeze_panes = "A7"
            sheet.auto_filter.ref = f"A6:{sheet.cell(sheet.max_row, sheet.max_column).coordinate}"
            for cell in sheet[6]:
                cell.fill = PatternFill("solid", fgColor="1F4E78")
                cell.font = Font(color="FFFFFF", bold=True)
            for column_index, header in enumerate(sheet[6], 1):
                name = str(header.value or "").casefold()
                for row_index in range(7, sheet.max_row + 1):
                    if "periodo" in name:
                        sheet.cell(row_index, column_index).number_format = "mm/yyyy"
                    elif name.startswith("fecha") or name in ("creado en", "actualizado en"):
                        sheet.cell(row_index, column_index).number_format = "dd/mm/yyyy"
            state_fills = {
                "verde": PatternFill("solid", fgColor="C6EFCE"),
                "amarillo": PatternFill("solid", fgColor="FFF2CC"),
                "rojo": PatternFill("solid", fgColor="FFC7CE"),
                "gris": PatternFill("solid", fgColor="E7E6E6"),
            }
            for row in sheet.iter_rows(min_row=7):
                for cell in row:
                    value = str(cell.value or "").casefold()
                    if value in ("activo", "completo", "pagado", "ok", "bajo", "al día", "recibido"):
                        cell.fill = state_fills["verde"]
                    elif value in ("pendiente", "revisar", "a revisar", "medio", "en proceso"):
                        cell.fill = state_fills["amarillo"]
                    elif value in ("vencido", "urgente", "alto", "con deuda", "baja"):
                        cell.fill = state_fills["rojo"]
                    elif value in ("no corresponde", "inactivo", "archivado"):
                        cell.fill = state_fills["gris"]
            for column in sheet.columns:
                width = min(max(len(str(cell.value or "")) for cell in column) + 2, 45)
                sheet.column_dimensions[column[0].column_letter].width = max(width, 12)
        workbook.save(destination)
        return destination

    def export_pdf(
        self, destination: Path, client_id: int, sections: list[str] | None = None
    ) -> Path:
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.lib.units import mm
            from reportlab.platypus import (
                PageBreak,
                Paragraph,
                SimpleDocTemplate,
                Spacer,
                Table,
                TableStyle,
            )
        except ImportError as error:
            raise RuntimeError(
                "Para exportar a PDF instalá las dependencias con: "
                "python -m pip install -r requirements.txt"
            ) from error

        sections = sections or list(self.SECTION_ORDER)
        destination.parent.mkdir(parents=True, exist_ok=True)
        styles = getSampleStyleSheet()
        story = []
        summary = self.ledger.summary(client_id)
        client = summary["client"]

        def pdf_value(key: str, value) -> str:
            from utils.formatters import display_date, display_period, number_ar
            if value is None: return ""
            name = str(key).casefold()
            if "periodo" in name: return display_period(str(value))
            if name.startswith("fecha") or name in ("creado_en", "actualizado_en"): return display_date(str(value))
            if any(term in name for term in ("importe", "saldo", "total")) and isinstance(value, (int, float)): return number_ar(value)
            return str(value)
        story.extend([
            Paragraph("Legajo del cliente", styles["Title"]),
            Paragraph(str(client["nombre_razon_social"]), styles["Heading2"]),
            Paragraph(f"Legajo: {client.get('legajo', '')}", styles["Normal"]),
            Paragraph(f"CUIT: {client['cuit_cuil']}", styles["Normal"]),
            Paragraph(f"Tipo de cliente: {summary['tipo_cliente']}", styles["Normal"]),
            Paragraph(f"Condición fiscal: {str(client['regimen_principal']).replace('_', ' ').title()}", styles["Normal"]),
            Paragraph(f"Estado: {summary['estado_cliente']}", styles["Normal"]),
            Paragraph(f"Fecha de exportación: {date.today().strftime('%d/%m/%Y')}", styles["Normal"]),
            Spacer(1, 8 * mm),
        ])
        for index, section in enumerate(sections):
            if index:
                story.append(PageBreak())
            story.append(Paragraph(self.section_title(section), styles["Heading1"]))
            rows = self.section_rows(client_id, section)
            if not rows:
                story.append(Paragraph("Sin información cargada", styles["Normal"]))
                continue
            if all(set(row) == {"Campo", "Valor"} for row in rows):
                data = [[Paragraph("Campo", styles["BodyText"]), Paragraph("Valor", styles["BodyText"])]]
                data.extend([
                    [Paragraph(str(row["Campo"]), styles["BodyText"]), Paragraph(pdf_value(str(row["Campo"]), row["Valor"]), styles["BodyText"])]
                    for row in rows
                ])
                table = Table(data, colWidths=(55 * mm, 125 * mm), repeatRows=1)
                table.setStyle(TableStyle([
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#D0D7DE")),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F8FA")]),
                    ("LEFTPADDING", (0, 0), (-1, -1), 6), ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                ]))
                story.append(table)
                continue
            keys = list(rows[0])
            if len(keys) <= 6 and all(list(row) == keys for row in rows):
                data = [[Paragraph(str(key).replace("_", " ").title(), styles["BodyText"]) for key in keys]]
                data.extend([
                    [Paragraph(pdf_value(key, row.get(key)), styles["BodyText"]) for key in keys]
                    for row in rows
                ])
                if keys == ["Sección", "Campo", "Valor"]:
                    widths = (45 * mm, 90 * mm, 45 * mm)
                else:
                    widths = tuple((180 / len(keys)) * mm for _key in keys)
                table = Table(data, colWidths=widths, repeatRows=1)
                table.setStyle(TableStyle([
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#D0D7DE")),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F8FA")]),
                    ("LEFTPADDING", (0, 0), (-1, -1), 4), ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                ]))
                story.append(table)
                continue
            for record_number, row in enumerate(rows, 1):
                data = [[Paragraph("Campo", styles["BodyText"]), Paragraph("Valor", styles["BodyText"])]]
                for key, value in row.items():
                    data.append([Paragraph(str(key).replace("_", " ").title(), styles["BodyText"]), Paragraph(pdf_value(str(key), value), styles["BodyText"])])
                table = Table(data, colWidths=(55 * mm, 125 * mm), repeatRows=1)
                table.setStyle(TableStyle([
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#D0D7DE")),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F8FA")]),
                    ("LEFTPADDING", (0, 0), (-1, -1), 6), ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                ]))
                story.append(table)
                if record_number < len(rows): story.append(Spacer(1, 4 * mm))

        def footer(canvas, document):
            canvas.saveState()
            canvas.setFont("Helvetica", 8)
            canvas.drawString(18 * mm, 10 * mm, f"{client['nombre_razon_social']} - {client['cuit_cuil']}")
            canvas.drawRightString(192 * mm, 10 * mm, f"Página {document.page}")
            canvas.restoreState()

        SimpleDocTemplate(
            str(destination), pagesize=A4, rightMargin=15 * mm, leftMargin=15 * mm,
            topMargin=16 * mm, bottomMargin=16 * mm,
        ).build(story, onFirstPage=footer, onLaterPages=footer)
        return destination

    def master_index_rows(self, client_ids: list[int]) -> list[dict]:
        result = []
        for client_id in client_ids:
            summary = self.ledger.summary(client_id)
            client = summary["client"]
            result.append({
                "Cliente": client["nombre_razon_social"],
                "Legajo": client.get("legajo", ""),
                "CUIT": client["cuit_cuil"],
                "Tipo": summary["tipo_cliente"],
                "Condición fiscal": str(client["regimen_principal"]).replace("_", " ").title(),
                "Estado cliente": summary["estado_cliente"],
                "Servicio": summary["servicio_contratado"],
                "Estado legajo": summary["estado_legajo"],
                "Pagos": summary["estado_pagos"],
                "Documentación": summary["estado_documentacion"],
                "Riesgo": summary["riesgo_general"],
                "Último control": summary["ultimo_control"],
                "Próximo vencimiento": summary["proximo_vencimiento"],
                "Observaciones": summary["observacion_ejecutiva"],
            })
        return result

    def export_master_index_excel(self, destination: Path, client_ids: list[int]) -> Path:
        rows = self.master_index_rows(client_ids)
        if not rows:
            raise ValueError("No hay clientes visibles para exportar.")
        destination.parent.mkdir(parents=True, exist_ok=True)
        pd.DataFrame(rows).to_excel(destination, index=False, sheet_name="Índice Clientes")
        workbook = load_workbook(destination)
        sheet = workbook.active
        sheet.freeze_panes = "A2"; sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.fill = PatternFill("solid", fgColor="1F4E78")
            cell.font = Font(color="FFFFFF", bold=True)
        for column in sheet.columns:
            sheet.column_dimensions[column[0].column_letter].width = min(max(len(str(cell.value or "")) for cell in column) + 2, 36)
        workbook.save(destination)
        return destination

    def export_master_index_pdf(self, destination: Path, client_ids: list[int]) -> Path:
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.lib.units import mm
            from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
        except ImportError as error:
            raise RuntimeError("Para exportar a PDF instalá: python -m pip install -r requirements.txt") from error
        rows = self.master_index_rows(client_ids)
        if not rows:
            raise ValueError("No hay clientes visibles para exportar.")
        destination.parent.mkdir(parents=True, exist_ok=True)
        styles = getSampleStyleSheet()
        headers = ("Cliente", "Legajo", "CUIT", "Condición fiscal", "Estado legajo", "Pagos", "Documentación", "Riesgo", "Próximo vencimiento")
        data = [[Paragraph(header, styles["BodyText"]) for header in headers]]
        data.extend([[Paragraph(str(row.get(header, "") or ""), styles["BodyText"]) for header in headers] for row in rows])
        table = Table(data, colWidths=(44*mm, 20*mm, 27*mm, 31*mm, 26*mm, 22*mm, 27*mm, 18*mm, 29*mm), repeatRows=1)
        table.setStyle(TableStyle([("BACKGROUND", (0,0), (-1,0), colors.HexColor("#1F4E78")), ("TEXTCOLOR", (0,0), (-1,0), colors.white), ("GRID", (0,0), (-1,-1), .25, colors.grey), ("VALIGN", (0,0), (-1,-1), "TOP"), ("FONTSIZE", (0,0), (-1,-1), 7)]))
        story = [Paragraph("Índice Maestro de Clientes", styles["Title"]), Paragraph(f"Fecha de exportación: {date.today().strftime('%d/%m/%Y')}", styles["Normal"]), Spacer(1, 5*mm), table]
        SimpleDocTemplate(str(destination), pagesize=landscape(A4), rightMargin=10*mm, leftMargin=10*mm, topMargin=12*mm, bottomMargin=12*mm).build(story)
        return destination

    def export_batch(
        self, destination: Path, client_ids: list[int], formats: tuple[str, ...] = ("xlsx", "pdf"),
        sections: list[str] | None = None,
    ) -> Path:
        if not client_ids:
            raise ValueError("No hay clientes seleccionados para exportar.")
        destination.parent.mkdir(parents=True, exist_ok=True)
        with tempfile.TemporaryDirectory() as directory:
            base = Path(directory)
            files = []
            for client_id in client_ids:
                summary = self.ledger.summary(client_id)
                client = summary["client"]
                stem = f"Legajo_Cliente_{self.safe_name(client['nombre_razon_social'])}_{client['cuit_cuil']}_{date.today().isoformat()}"
                if "xlsx" in formats:
                    files.append(self.export_excel(base / f"{stem}.xlsx", client_id, sections))
                if "pdf" in formats:
                    files.append(self.export_pdf(base / f"{stem}.pdf", client_id, sections))
            with zipfile.ZipFile(destination, "w", zipfile.ZIP_DEFLATED) as archive:
                for path in files:
                    archive.write(path, path.name)
        return destination
