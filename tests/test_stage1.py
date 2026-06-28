from __future__ import annotations

import sqlite3
import tempfile
import unittest
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

from database import Database, initialize_database
from database.seed import seed_reference_data
from models import Client, FiscalProfile, MonotributoProfile, Voucher
from services import (
    AdministrativeService,
    ClientService,
    AlertService,
    ConfigService,
    ImportService,
    IibbService,
    LedgerExportService,
    LedgerService,
    MonotributoService,
    ReportService,
    VoucherService,
)


class StageOneTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temporary = tempfile.TemporaryDirectory()
        self.path = Path(self.temporary.name)
        self.database = Database(self.path / "test.db")
        initialize_database(self.database)
        seed_reference_data(self.database)
        self.config = ConfigService(self.database)
        self.clients = ClientService(self.database)
        self.vouchers = VoucherService(self.database, self.config)
        self.mono = MonotributoService(self.database, self.vouchers, self.config)
        self.client_id = self.clients.save(
            Client("Persona de Prueba", "20123456786"),
            FiscalProfile(regimen_principal="monotributista"),
            MonotributoProfile(categoria_actual="A"),
        )

    def tearDown(self) -> None:
        self.temporary.cleanup()

    def test_schema_contains_all_requested_tables(self) -> None:
        rows = self.database.query(
            "SELECT name FROM sqlite_master WHERE type = 'table' AND name NOT LIKE 'sqlite_%'"
        )
        names = {row["name"] for row in rows}
        expected = {
            "clientes",
            "datos_fiscales_cliente",
            "obligaciones_fiscales",
            "cliente_obligaciones",
            "monotributo_cliente",
            "comprobantes_ventas",
            "comprobantes_compras",
            "iibb_monotributo",
            "categorias_monotributo",
            "recategorizaciones_monotributo",
            "alertas_fiscales",
            "documentacion",
            "tareas",
            "vencimientos",
            "honorarios",
            "configuracion",
            "configuracion_alertas_cliente",
            "cliente_legajo_campos",
            "cliente_legajo_registros",
            "cliente_historial",
        }
        self.assertTrue(expected.issubset(names))

    def test_client_duplicate_and_logical_deactivation(self) -> None:
        with self.assertRaisesRegex(ValueError, "CUIT/CUIL"):
            self.clients.save(
                Client("Duplicado", "20123456786"), FiscalProfile(), None
            )
        self.clients.deactivate(self.client_id)
        self.assertEqual(self.clients.list_clients(), [])
        self.assertEqual(len(self.clients.list_clients(include_inactive=True)), 1)

    def test_alert_recalculation_does_not_require_monotributo_profile(self) -> None:
        other_id = self.clients.save(
            Client("Cliente sin régimen", "20333444559"),
            FiscalProfile(regimen_principal="sin_definir"),
            None,
        )
        count = AlertService(
            self.database, self.config, self.vouchers, self.mono
        ).refresh(other_id)
        self.assertGreaterEqual(count, 1)

    def test_permanent_client_delete_cascades_to_vouchers(self) -> None:
        period = date.today().strftime("%Y-%m")
        self.vouchers.create(
            "ventas",
            Voucher(
                cliente_id=self.client_id,
                fecha=date.today().isoformat(),
                periodo_fiscal=period,
                tipo_comprobante="Factura C",
                punto_venta="99",
                numero_comprobante="1",
                contraparte_nombre="Comprador",
                importe_original=1000,
            ),
        )
        self.database.execute(
            """
            INSERT INTO alertas_fiscales(
                cliente_id, periodo, tipo_alerta, descripcion, gravedad
            ) VALUES (?, ?, 'prueba', 'Alerta de prueba', 'media')
            """,
            (self.client_id, period),
        )
        deleted = self.clients.delete_permanently(self.client_id)
        self.assertEqual(deleted["ventas"], 1)
        self.assertIsNone(
            self.database.query_one("SELECT id FROM clientes WHERE id = ?", (self.client_id,))
        )
        self.assertEqual(
            self.database.query_one(
                "SELECT COUNT(*) AS n FROM comprobantes_ventas WHERE cliente_id = ?",
                (self.client_id,),
            )["n"],
            0,
        )
        self.assertEqual(
            self.database.query_one(
                "SELECT COUNT(*) AS n FROM alertas_fiscales WHERE cliente_id = ?",
                (self.client_id,),
            )["n"],
            0,
        )

    def test_fiscal_sign_usd_conversion_and_annulled(self) -> None:
        period = date.today().strftime("%Y-%m")
        normal = Voucher(
            cliente_id=self.client_id,
            fecha=date.today().isoformat(),
            periodo_fiscal=period,
            tipo_comprobante="Factura C",
            punto_venta="1",
            numero_comprobante="1",
            contraparte_nombre="Cliente A",
            moneda="USD",
            tipo_cambio=1000,
            importe_original=600,
        )
        credit = Voucher(
            cliente_id=self.client_id,
            fecha=date.today().isoformat(),
            periodo_fiscal=period,
            tipo_comprobante="Nota de Crédito C",
            punto_venta="1",
            numero_comprobante="2",
            contraparte_nombre="Cliente A",
            importe_original=100_000,
        )
        annulled = Voucher(
            cliente_id=self.client_id,
            fecha=date.today().isoformat(),
            periodo_fiscal=period,
            tipo_comprobante="Factura C",
            punto_venta="1",
            numero_comprobante="3",
            contraparte_nombre="Cliente B",
            importe_original=200_000,
            estado="anulado",
        )
        self.vouchers.create("ventas", normal)
        self.vouchers.create("ventas", credit)
        self.vouchers.create("ventas", annulled)
        self.vouchers.create(
            "ventas",
            Voucher(
                cliente_id=self.client_id,
                fecha=date.today().isoformat(),
                periodo_fiscal=period,
                tipo_comprobante="Factura C",
                punto_venta="1",
                numero_comprobante="4",
                contraparte_nombre="",
                importe_original=500,
            ),
        )
        rows = self.vouchers.list("ventas", self.client_id, period)
        by_number = {row["numero_comprobante"]: row for row in rows}
        self.assertEqual(by_number["1"]["importe_neto_fiscal"], 600_000)
        self.assertEqual(by_number["2"]["importe_neto_fiscal"], -100_000)
        self.assertEqual(by_number["3"]["importe_neto_fiscal"], 0)
        self.assertEqual(by_number["4"]["contraparte_nombre"], "CONSUMIDOR FINAL")
        self.assertEqual(self.vouchers.stats("ventas", self.client_id)["mes"], 500_500)
        alert_count = self.database.query_one(
            "SELECT COUNT(*) AS n FROM alertas_fiscales WHERE cliente_id = ?",
            (self.client_id,),
        )
        self.assertEqual(alert_count["n"], 2)
        deleted = self.vouchers.delete_selected(
            "ventas",
            self.client_id,
            [by_number["1"]["id"], by_number["2"]["id"]],
        )
        self.assertEqual(deleted, 2)
        remaining = self.vouchers.list("ventas", self.client_id, period)
        self.assertEqual(
            [row["numero_comprobante"] for row in remaining], ["4", "3"]
        )

    def test_dashboard_category_and_excel_export(self) -> None:
        period = date.today().strftime("%Y-%m")
        self.vouchers.create(
            "ventas",
            Voucher(
                cliente_id=self.client_id,
                fecha=date.today().isoformat(),
                periodo_fiscal=period,
                tipo_comprobante="Factura C",
                punto_venta="2",
                numero_comprobante="10",
                contraparte_nombre="Comprador",
                importe_original=1_000_000,
            ),
        )
        dashboard = self.mono.dashboard(self.client_id)
        self.assertEqual(dashboard["sales"]["mes"], 1_000_000)
        self.assertEqual(dashboard["suggested_category"], "A")

        output = self.path / "ventas.xlsx"
        ReportService(self.vouchers).export_vouchers(
            "ventas", output, self.client_id, period
        )
        workbook = load_workbook(output)
        self.assertEqual(workbook.sheetnames, ["Ventas"])
        self.assertEqual(workbook["Ventas"]["A1"].value, "Cliente del estudio")

    def test_each_accounting_section_can_be_exported_to_excel(self) -> None:
        period = date.today().strftime("%Y-%m")
        self.vouchers.create(
            "ventas",
            Voucher(
                cliente_id=self.client_id,
                fecha=date.today().isoformat(),
                periodo_fiscal=period,
                tipo_comprobante="Factura C",
                punto_venta="3",
                numero_comprobante="20",
                contraparte_nombre="Cliente del exterior",
                contraparte_documento="",
                moneda="USD",
                tipo_cambio=1000,
                importe_original=600,
            ),
        )
        exporter = ReportService(self.vouchers)
        expected = {
            "resumen_mensual": ("Resumen mensual", "Período"),
            "significativos": ("Significativos", "Fecha"),
            "moneda_extranjera": ("Moneda extranjera", "Fecha"),
            "ranking": ("Ranking", "Puesto"),
        }
        for section, (sheet_name, first_header) in expected.items():
            output = self.path / f"{section}.xlsx"
            exporter.export_accounting_section(
                "ventas", section, output, self.client_id
            )
            workbook = load_workbook(output)
            self.assertEqual(workbook.sheetnames, [sheet_name])
            sheet = workbook[sheet_name]
            self.assertEqual(sheet["A1"].value, first_header)
            self.assertEqual(sheet.freeze_panes, "A2")
            self.assertGreaterEqual(sheet.max_row, 2)
            workbook.close()

    def test_last_twelve_months_totals_zero_division_and_excel(self) -> None:
        iibb = IibbService(self.database, self.config)
        iibb.save_profile(
            self.client_id,
            {"regimen_principal": "Régimen general/local", "alicuota": 0.035},
        )
        self.vouchers.create(
            "ventas",
            Voucher(
                cliente_id=self.client_id,
                fecha="2026-06-10",
                periodo_fiscal="2026-06",
                tipo_comprobante="Factura C",
                punto_venta="4",
                numero_comprobante="1",
                contraparte_nombre="Cliente junio",
                importe_original=1000,
            ),
        )
        self.vouchers.create(
            "compras",
            Voucher(
                cliente_id=self.client_id,
                fecha="2026-06-11",
                periodo_fiscal="2026-06",
                tipo_comprobante="Factura A",
                punto_venta="4",
                numero_comprobante="2",
                contraparte_nombre="Proveedor junio",
                importe_original=250,
            ),
        )
        self.vouchers.create(
            "compras",
            Voucher(
                cliente_id=self.client_id,
                fecha="2026-05-11",
                periodo_fiscal="2026-05",
                tipo_comprobante="Factura A",
                punto_venta="4",
                numero_comprobante="3",
                contraparte_nombre="Proveedor mayo",
                importe_original=100,
            ),
        )
        iibb.save_fixed_amount(self.client_id, "2026-06", 50)
        reports = ReportService(self.vouchers)
        report = reports.last_twelve_months(
            self.client_id, reference=date(2026, 6, 27)
        )
        self.assertEqual(len(report["rows"]), 12)
        self.assertEqual(report["rows"][0]["periodo"], "2025-07")
        self.assertEqual(report["rows"][-1]["periodo"], "2026-06")
        may = next(row for row in report["rows"] if row["periodo"] == "2026-05")
        self.assertEqual(may["porcentaje_compras"], 0)
        self.assertEqual(report["totals"]["ventas"], 1000)
        self.assertEqual(report["totals"]["compras"], 350)
        self.assertEqual(report["totals"]["resultado"], 650)
        self.assertAlmostEqual(report["totals"]["porcentaje_compras"], 0.35)
        self.assertEqual(report["totals"]["ingresos_brutos"], 35)
        self.assertEqual(report["totals"]["regimen_simplificado"], 50)
        custom = reports.last_twelve_months(
            self.client_id,
            date_from="2026-05-01",
            date_to="2026-06-30",
        )
        self.assertEqual([row["mes"] for row in custom["rows"]], ["05-2026", "06-2026"])

        output = self.path / "ultimos_12.xlsx"
        reports.export_last_twelve_months(
            output, self.client_id, reference=date(2026, 6, 27)
        )
        workbook = load_workbook(output)
        sheet = workbook["Últimos 12 meses"]
        self.assertEqual(sheet.max_row, 14)
        self.assertEqual(sheet["A14"].value, "TOTAL")
        self.assertEqual(sheet["B14"].value, 1000)
        self.assertEqual(sheet["E14"].value, 0.35)
        self.assertEqual(sheet["E14"].number_format, "0.0%")
        self.assertIsInstance(sheet["A2"].value, date)
        self.assertEqual(sheet["A2"].number_format, "mm-yyyy")
        workbook.close()

        monthly_output = self.path / "ventas_rango.xlsx"
        reports.export_named(
            "ventas_mensuales",
            monthly_output,
            self.client_id,
            "2026-06-01",
            "2026-06-30",
        )
        workbook = load_workbook(monthly_output)
        sheet = workbook["Reporte"]
        self.assertIsInstance(sheet["A2"].value, date)
        self.assertEqual(sheet["A2"].number_format, "mm-yyyy")
        workbook.close()

    def test_activity_code_and_client_alert_configuration(self) -> None:
        configured_id = self.clients.save(
            Client("Actividad configurada", "20333444559"),
            FiscalProfile(regimen_principal="monotributista"),
            MonotributoProfile(
                categoria_actual="B",
                actividad_fiscal="Servicios",
                codigo_actividad="620100",
            ),
        )
        bundle = self.clients.get_bundle(configured_id)
        self.assertEqual(bundle["monotributo"]["codigo_actividad"], "620100")
        with self.assertRaisesRegex(ValueError, "solamente números"):
            self.clients.save(
                Client("Actividad inválida", "20444555661"),
                FiscalProfile(regimen_principal="monotributista"),
                MonotributoProfile(codigo_actividad="ABC-1"),
            )

        self.config.save_client_alerts(
            configured_id,
            {
                "monto_comprobante_significativo": 2_000_000,
                "monotributo_alerta_porcentaje": 0.75,
            },
        )
        self.assertEqual(
            self.config.get_client_float(
                configured_id, "monto_comprobante_significativo"
            ),
            2_000_000,
        )
        self.vouchers.create(
            "ventas",
            Voucher(
                cliente_id=configured_id,
                fecha="2026-06-12",
                periodo_fiscal="2026-06",
                tipo_comprobante="Factura C",
                punto_venta="5",
                numero_comprobante="1",
                contraparte_nombre="Sin alerta elevada",
                importe_original=1_000_000,
            ),
        )
        count = self.database.query_one(
            """SELECT COUNT(*) n FROM alertas_fiscales
               WHERE cliente_id=? AND tipo_alerta='venta_significativa'""",
            (configured_id,),
        )
        self.assertEqual(count["n"], 0)

        output = self.path / "legajo_cliente.xlsx"
        ReportService(self.vouchers).export_client_file(output, configured_id)
        workbook = load_workbook(output)
        self.assertIn("Resumen", workbook.sheetnames)
        self.assertIn("Monotributo", workbook.sheetnames)
        headers = [cell.value for cell in workbook["Monotributo"][1]]
        self.assertIn("codigo_actividad", headers)
        workbook.close()

    def test_iibb_monthly_detail_and_payable(self) -> None:
        iibb = IibbService(self.database, self.config)
        iibb.save_profile(
            self.client_id,
            {"regimen_principal": "Régimen general/local", "alicuota": 0.035},
        )
        for number, voucher_type, amount in (
            ("31", "Factura C", 1000),
            ("32", "Nota de Crédito C", 200),
        ):
            self.vouchers.create(
                "ventas",
                Voucher(
                    cliente_id=self.client_id,
                    fecha="2026-06-15",
                    periodo_fiscal="2026-06",
                    tipo_comprobante=voucher_type,
                    punto_venta="6",
                    numero_comprobante=number,
                    contraparte_nombre="Cliente IIBB",
                    importe_original=amount,
                ),
            )
        result = iibb.calculate_and_save(
            self.client_id, "2026-06", retentions=10
        )
        detail = iibb.monthly_detail(self.client_id, "2026-06")
        self.assertEqual(len(detail["rows"]), 2)
        self.assertEqual(detail["base"], 800)
        self.assertEqual(result["determined"], 28)
        self.assertEqual(detail["retentions"], 10)
        self.assertEqual(detail["payable"], 18)

    def test_tasks_and_fees_can_be_updated_and_deleted(self) -> None:
        service = AdministrativeService(self.database)
        task_id = service.create(
            "tareas",
            {
                "cliente_id": self.client_id,
                "modulo": "Monotributo",
                "periodo": "2026-06",
                "titulo": "Presentar DDJJ",
                "estado": "pendiente",
                "prioridad": "media",
            },
        )
        task = service.get("tareas", task_id)
        task.update({"titulo": "Presentar DDJJ corregida", "prioridad": "alta"})
        service.update("tareas", task_id, task)
        self.assertEqual(service.get("tareas", task_id)["prioridad"], "alta")
        self.assertEqual(service.delete("tareas", task_id), 1)
        self.assertIsNone(service.get("tareas", task_id))

        fee_id = service.create(
            "honorarios",
            {
                "cliente_id": self.client_id,
                "servicio": "Liquidación mensual",
                "periodo": "2026-06",
                "importe": "10000",
                "estado": "pendiente de facturar",
                "saldo_pendiente": "10000",
            },
        )
        fee = service.get("honorarios", fee_id)
        fee.update({"importe": "12000", "saldo_pendiente": "5000"})
        service.update("honorarios", fee_id, fee)
        self.assertEqual(service.get("honorarios", fee_id)["importe"], 12000)
        self.assertEqual(service.delete("honorarios", fee_id), 1)

        due_id = service.create(
            "vencimientos",
            {
                "cliente_id": self.client_id,
                "impuesto": "Ingresos Brutos",
                "periodo": "2026-06",
                "fecha_vencimiento": "2026-07-15",
                "estado": "pendiente",
            },
        )
        due = service.get("vencimientos", due_id)
        due.update({"estado": "pagado", "fecha_vencimiento": "2026-07-16"})
        service.update("vencimientos", due_id, due)
        self.assertEqual(service.get("vencimientos", due_id)["estado"], "pagado")
        self.assertEqual(service.delete("vencimientos", due_id), 1)

    def test_supplier_matrix_by_month_and_total(self) -> None:
        purchases = (
            ("2026-01-10", "1", "Proveedor Uno", 100),
            ("2026-02-10", "2", "Proveedor Uno", 200),
            ("2026-01-12", "3", "Proveedor Dos", 50),
        )
        for voucher_date, number, provider, amount in purchases:
            self.vouchers.create(
                "compras",
                Voucher(
                    cliente_id=self.client_id,
                    fecha=voucher_date,
                    periodo_fiscal=voucher_date[:7],
                    tipo_comprobante="Factura A",
                    punto_venta="30",
                    numero_comprobante=number,
                    contraparte_nombre=provider,
                    importe_original=amount,
                ),
            )
        reports = ReportService(self.vouchers)
        matrix = reports.supplier_matrix(self.client_id, 2026)
        self.assertEqual(
            list(matrix.columns),
            [
                "Proveedor", "Enero", "Febrero", "Marzo", "Abril", "Mayo",
                "Junio", "Julio", "Agosto", "Septiembre", "Octubre",
                "Noviembre", "Diciembre", "Total",
            ],
        )
        provider = matrix[matrix["Proveedor"] == "Proveedor Uno"].iloc[0]
        self.assertEqual(provider["Enero"], 100)
        self.assertEqual(provider["Febrero"], 200)
        self.assertEqual(provider["Total"], 300)
        self.assertEqual(matrix.iloc[-1]["Total"], 350)

        output = self.path / "proveedores.xlsx"
        reports.export_supplier_matrix(output, self.client_id, 2026)
        workbook = load_workbook(output)
        self.assertEqual(workbook.sheetnames, ["Proveedores 2026"])
        self.assertEqual(workbook["Proveedores 2026"]["A1"].value, "Proveedor")
        workbook.close()

    def test_integral_ledger_all_sections_summary_and_exports(self) -> None:
        ledger = LedgerService(self.database)
        payment_id = ledger.save_record(
            self.client_id,
            "pagos",
            {
                "fecha_emision": "01/06/2026",
                "periodo": "06-2026",
                "concepto": "Abono mensual",
                "importe_facturado": "10000",
                "importe_cobrado": "4000",
                "estado_pago": "Pago parcial",
                "fecha_vencimiento": "10/06/2026",
                "responsable": "NATALIA",
            },
        )
        self.assertEqual(ledger.get_record(payment_id)["saldo"], 6000)
        for section in ledger.SECTIONS:
            if section == "pagos":
                continue
            ledger.save_record(
                self.client_id,
                section,
                {
                    "descripcion": f"Registro {section}",
                    "estado": "Recibido" if section == "documentacion" else "pendiente",
                    "responsable": "NATALIA",
                },
            )
        summary = ledger.summary(self.client_id)
        self.assertEqual(summary["estado_legajo"], "Completo")
        self.assertEqual(summary["pagos_pendientes"], 1)
        self.assertEqual(summary["total_pendiente"], 6000)
        self.assertGreaterEqual(len(ledger.history(self.client_id)), len(ledger.SECTIONS))

        exporter = LedgerExportService(self.database, ledger)
        xlsx = self.path / "legajo_integral.xlsx"
        pdf = self.path / "legajo_integral.pdf"
        archive = self.path / "legajos.zip"
        index_xlsx = self.path / "indice_clientes.xlsx"
        index_pdf = self.path / "indice_clientes.pdf"
        exporter.export_excel(xlsx, self.client_id)
        exporter.export_pdf(pdf, self.client_id)
        exporter.export_batch(archive, [self.client_id], ("xlsx", "pdf"))
        exporter.export_master_index_excel(index_xlsx, [self.client_id])
        exporter.export_master_index_pdf(index_pdf, [self.client_id])
        workbook = load_workbook(xlsx)
        self.assertIn("Resumen", workbook.sheetnames)
        self.assertIn("Pagos del Cliente", workbook.sheetnames)
        self.assertIn("Riesgos", workbook.sheetnames)
        self.assertEqual(workbook["Resumen"]["A1"].value, "Cliente")
        self.assertEqual(workbook["Resumen"]["A6"].value, "Campo")
        workbook.close()
        self.assertGreater(pdf.stat().st_size, 1000)
        self.assertGreater(archive.stat().st_size, 1000)
        self.assertGreater(index_xlsx.stat().st_size, 1000)
        self.assertGreater(index_pdf.stat().st_size, 1000)
        self.assertEqual(ledger.delete_record(self.client_id, payment_id), 1)

    def test_arca_csv_import_and_iibb(self) -> None:
        source = self.path / "emitidos.csv"
        source.write_text(
            '"Fecha de Emisión";"Tipo de Comprobante";"Punto de Venta";'
            '"Número Desde";"Número Hasta";"Cód. Autorización";'
            '"Tipo Doc. Receptor";"Nro. Doc. Receptor";"Denominación Receptor";'
            '"Tipo Cambio";"Moneda";"Imp. Total"\n'
            '2026-06-01;11;1;10;10;123;80;30700000001;Cliente Uno;1;$;62150,30\n'
            '2026-06-02;13;1;11;11;124;80;30700000001;Cliente Uno;1;$;1250000,50\n',
            encoding="utf-8",
        )
        importer = ImportService(self.database, self.vouchers)
        preview = importer.preview(source, "ventas")
        self.assertEqual(preview.missing, [])
        result = importer.import_rows(preview, self.client_id)
        self.assertEqual(result["imported"], 2)
        rows = self.vouchers.list("ventas", self.client_id, "2026-06")
        self.assertEqual(rows[0]["importe_neto_fiscal"], -1_250_000.50)
        self.assertEqual(rows[0]["nombre_archivo_origen"], "emitidos.csv")
        self.assertIsNotNone(rows[0]["id_importacion"])

        iibb = IibbService(self.database, self.config)
        iibb.save_profile(
            self.client_id,
            {"regimen_principal": "ARBA REG GENERAL", "alicuota": 0.035},
        )
        calculation = iibb.calculate_and_save(self.client_id, "2026-06")
        self.assertEqual(calculation["base"], 0)
        deleted = importer.delete_batch(self.client_id, result["import_id"])
        self.assertEqual(deleted, 2)
        self.assertEqual(self.vouchers.list("ventas", self.client_id, "2026-06"), [])

        no_receiver = self.path / "emitidos_sin_receptor.csv"
        no_receiver.write_text(
            '"Fecha";"Tipo de Comprobante";"Punto de Venta";'
            '"Número Desde";"Tipo Cambio";"Moneda";"Imp. Total"\n'
            '2026-06-03;11;2;20;1;$;500,00\n',
            encoding="utf-8",
        )
        preview = importer.preview(no_receiver, "ventas")
        self.assertEqual(preview.missing, [])
        result = importer.import_rows(preview, self.client_id)
        self.assertEqual(result["imported"], 1)
        row = self.vouchers.list("ventas", self.client_id, "2026-06")[0]
        self.assertEqual(row["contraparte_nombre"], "CONSUMIDOR FINAL")
        importer.delete_batch(self.client_id, result["import_id"])


if __name__ == "__main__":
    unittest.main()
