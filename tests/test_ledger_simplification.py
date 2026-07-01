from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from database import Database, initialize_database
from database.seed import seed_reference_data
from models import Client, FiscalProfile, MonotributoProfile
from services import ClientService, LedgerExportService, LedgerService


class LedgerSimplificationTests(unittest.TestCase):
    REMOVED = {
        "obligaciones",
        "contactos_arca",
        "domicilios_arca",
        "migratorios_arca",
        "societario",
        "baja_historial",
        "historial",
    }

    def setUp(self) -> None:
        self.temporary = tempfile.TemporaryDirectory()
        self.database = Database(Path(self.temporary.name) / "test.db")
        initialize_database(self.database)
        seed_reference_data(self.database)
        self.client_id = ClientService(self.database).save(
            Client("Cliente Manual", "20333444556", rubro="Servicios"),
            FiscalProfile(regimen_principal="monotributista"),
            MonotributoProfile(categoria_actual="A"),
        )
        self.ledger = LedgerService(self.database)
        self.exporter = LedgerExportService(self.database, self.ledger)

    def tearDown(self) -> None:
        self.temporary.cleanup()

    def test_only_active_sections_are_visible_and_exportable(self) -> None:
        self.assertEqual(len(self.ledger.VISIBLE_SECTIONS) + 2, 14)
        self.assertTrue(self.REMOVED.isdisjoint(self.ledger.SECTIONS))
        self.assertTrue(self.REMOVED.isdisjoint(self.exporter.SECTION_ORDER))
        with self.assertRaisesRegex(ValueError, "legajo activo"):
            self.exporter.section_rows(self.client_id, "obligaciones")

    def test_summary_and_full_excel_ignore_old_removed_records(self) -> None:
        self.database.execute(
            """INSERT INTO cliente_legajo_registros
               (cliente_id,seccion,descripcion,estado,importe,saldo,vencimiento,datos_json)
               VALUES(?,?,?,?,?,?,?,?)""",
            (
                self.client_id,
                "obligaciones",
                "Registro antiguo",
                "vencido",
                99999,
                99999,
                "2020-01-01",
                '{"tipo_obligacion":"IVA"}',
            ),
        )
        self.ledger.save_record(
            self.client_id,
            "datos_complementarios",
            {
                "domicilio_real": "Calle Manual 123",
                "contacto_principal": "Persona de contacto",
                "tipo_societario": "Dato antiguo que no debe mostrarse",
            },
        )
        summary = self.ledger.summary(self.client_id)
        self.assertNotIn("obligaciones_pendientes", summary)
        self.assertNotIn("obligaciones_vencidas", summary)
        self.assertNotIn("Societario", summary["estados_area"])
        self.assertEqual(summary["proximo_vencimiento"], "—")

        data_rows = self.exporter.section_rows(self.client_id, "datos_cliente")
        labels = {row["Campo"] for row in data_rows}
        self.assertIn("Domicilio real", labels)
        self.assertIn("Contacto principal", labels)
        self.assertNotIn("Tipo societario", labels)

        output = Path(self.temporary.name) / "legajo_activo.xlsx"
        self.exporter.export_excel(output, self.client_id)
        workbook = load_workbook(output, read_only=True)
        self.assertEqual(len(workbook.sheetnames), 15)
        self.assertIn("Responsable Inscripto", workbook.sheetnames)
        self.assertNotIn("Valores Mensuales", workbook.sheetnames)
        self.assertNotIn("Contactos ARCA", workbook.sheetnames)
        self.assertNotIn("Domicilios ARCA", workbook.sheetnames)
        self.assertNotIn("Datos Migratorios", workbook.sheetnames)
        self.assertNotIn("Societario Libros", workbook.sheetnames)
        self.assertNotIn("Historial", workbook.sheetnames)
        workbook.close()


if __name__ == "__main__":
    unittest.main()
